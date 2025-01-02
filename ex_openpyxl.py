import os
import pandas as pd
from pathlib import Path
from openpyxl.styles import Alignment,Font
from openpyxl import load_workbook,Workbook

if __name__ == '__main__':

    cur_dirname = os.path.dirname(__file__)
    input_path = os.path.join(cur_dirname, 'input.xlsx')

    df = pd.read_excel(input_path)

    print(df)

    # 加载需要更新的的工作簿
    wb = load_workbook(input_path, data_only=True)
    # 获取指定的工作表
    ws = wb['Sheet1']
    for row in ws['A1:K11']:
        for cell in row:
            print(cell.value, end=' ')
        print()
    # 关闭工作簿
    wb.close()

    # 加载需要更新的的工作簿
    wb = load_workbook(input_path)
    # 获取指定的工作表
    ws = wb['Sheet1']
    data = [100]*10
    for i in range(10):
        ws.cell(row=i+2, column=1, value=data[i])
    # 保存工作簿
    wb.save(input_path)
    # 关闭工作簿
    wb.close()

    print('当前目录：', Path.cwd())
    
    dest_path = os.path.dirname(__file__)
    file_name = '全市编制年统数.xlsx'
    file_path = os.path.join(dest_path, file_name)

    workbook = load_workbook(filename = file_path) 
    print('表格列表：', workbook.sheetnames)

    sheet = workbook.active
    #sheet = workbook['2023年度']
    print('活动表格：', sheet)
    print('表格尺寸：', sheet.dimensions)
    print('表格行数：', len(list(sheet.rows)))
    print('表格列数：', len(list(sheet.columns)))

    #cell = sheet['E4']
    #print(f'{cell.coordinate}(R{cell.row}C{cell.column}({cell.column_letter}))={cell.value}') 
    #cell = sheet.cell(row=25,column=2)
    #print(f'{cell.coordinate}(R{cell.row}C{cell.column}({cell.column_letter}))={cell.value}') 

    # 获取 A1:C2 区域的值
    cells = sheet['B5:D6']
    for i in cells: 
        for j in i:
            pass

    # 按行获取值
    for i in sheet.iter_rows(min_row=5, max_row=6, min_col=2, max_col=4):
        for j in i: 
            pass

    # 按列获取值
    for i in sheet.iter_cols(min_row=5, max_row=6, min_col=2, max_col=4):
        for j in i:
            pass

    # 获取所有行
    for i in sheet.rows: 
        pass

    # .append()方式：会在表格已有的数据后面，按行插入；
    # 这个操作很有用，爬虫得到的数据，可以使用该方式保存成 Excel 文件；
    data = [
        ["唐僧","男","180cm"], 
        ["孙悟空","男","188cm"], 
        ["猪八戒","男","175cm"], 
        ["沙僧","男","176cm"], 
        ] 
    for i in data: 
        sheet.append(i)

    # 在 python 中插入 excel 公式
    #for i in range(2,16):
    #    sheet[f'D{i}'] = '=1+2+3+4'

    # .insert_cols(idx=数字编号, amount=要插入的列数)，插入的位置是在 idx 列数的左侧插入；
    # .insert_rows(idx=数字编号, amount=要插入的行数)，插入的行数是在 idx 行数的上方插入；
    #sheet.insert_cols(idx=4,amount=2)
    #sheet.insert_rows(idx=5,amount=4)

    # .delete_rows(idx=数字编号, amount=要删除的行数)
    # .delete_cols(idx=数字编号, amount=要删除的列数)
    #sheet.delete_cols(idx=2,amount=2) 
    #sheet.delete_rows(idx=3,amount=2)

    # .move_range("数据区域",rows=,cols=)：正整数为向下或向右、负整数为向左或向上；
    # 向左移动两列，向下移动两行
    #sheet.move_range("C1:D4",rows=2,cols=-1)

    # .create_sheet("新的 sheet 名")：创建一个新的 sheet 表；
    workbook.create_sheet("我是一个新的 sheet") 
    print(workbook.sheetnames) 

    # .remove("sheet 名")：删除某个 sheet 表；
    sheet = workbook['Sheet1']
    workbook.remove(sheet)
    print(workbook.sheetnames)

    # 这个操作的实质，就是复制某个 excel 表中的 sheet 表，然后将文件存储到另外一张excel 表中；
    sheet = workbook['县区名单'] 
    workbook.copy_worksheet(sheet) 
    print(workbook.sheetnames)

    sheet = workbook['我是一个新的 sheet'] 
    sheet.title = "我是修改后的 sheet 名"
    print(workbook.sheetnames)

    file_path2 = os.path.join(dest_path, 'R.xlsx')
    workbook.save(filename=file_path2)
    workbook.close()

    file_path2 = os.path.join(dest_path, 'R2.xlsx')
    workbook = Workbook() 
    sheet = workbook.active 
    sheet.title = "表格 1"
    # sheet.freeze_panes='单元格'：冻结窗口；
    # 冻结窗口以后，你可以打开源文件，进行检验；
    sheet.freeze_panes = "C3" 

    # sheet.auto_filter.ref：给表格添加“筛选器” 
    # .auto_filter.ref = sheet.dimension 给所有字段添加筛选器；
    # .auto_filter.ref = "A1" 给 A1 这个格子添加“筛选器”，就是给第一列添加“筛选器”；
    # sheet.auto_filter.ref = sheet["A1"] 

    # 修改字体样式 
    # Font(name=字体名称,size=字体大小,bold=是否加粗,italic=是否斜体,color=字体颜色)
    cell = sheet["A1"] 
    font = Font(name="微软雅黑",size=20,bold=True,italic=True,color="FF0000")
    cell.font = font
    sheet["A1"] = "哈喽"

    # 获取表格中格子的字体样式
    cell = sheet["A1"] 
    font = cell.font 
    print(font.name, font.size, font.bold, font.italic, font.color)

    # 设置对齐样式 
    # Alignment(horizontal=水平对齐模式,vertical=垂直对齐模式,text_rotation=旋转角度,wrap_text=是否自动换行)
    # 水平对齐："general", "left", "center", "right", "fill", "justify", "centerContinuous", "distributed"；
    # 垂直对齐："top", "center", "bottom", "justify", "distributed"；
    cell = sheet["A1"] 
    alignment = Alignment(horizontal="center",vertical="center",text_rotation=45,wrap_text=True)
    cell.alignment = alignment 

    workbook.save(filename = file_path2)
    workbook.close()